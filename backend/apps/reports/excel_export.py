"""
reports/excel_export.py
Excel (XLSX) export for all major reports using openpyxl.
Falls back to CSV if openpyxl not installed.
"""
import io
from django.http import HttpResponse
from django.utils import timezone


def _workbook():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return openpyxl, True
    except ImportError:
        return None, False


def _xl_response(filename: str):
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _style_header(ws, row, openpyxl):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill(start_color='0D1017', end_color='0D1017', fill_type='solid')
    font = Font(color='FFFFFF', bold=True, size=10)
    for cell in ws[row]:
        cell.font      = font
        cell.fill      = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


def export_loan_portfolio(queryset=None) -> HttpResponse:
    """Export full loan portfolio to Excel."""
    from apps.loans.models import Loan

    qs = queryset or Loan.objects.select_related(
        'customer', 'product', 'branch', 'loan_officer'
    ).order_by('-created_at')

    oxl, has_xl = _workbook()
    if not has_xl:
        # Fallback CSV
        from apps.customers.bulk import export_loans_csv
        return export_loans_csv(qs)

    wb  = oxl.Workbook()
    ws  = wb.active
    ws.title = 'Loan Portfolio'
    ws.freeze_panes = 'A2'

    headers = ['Loan ID','Customer Name','Customer UID','Product','Branch','Loan Officer',
               'Principal','Interest Rate','Total Amount','Total Paid','Balance',
               'Status','Disbursed Date','Due Date','Days Overdue']
    ws.append(headers)
    _style_header(ws, 1, oxl)

    today = timezone.now().date()
    for loan in qs.iterator(chunk_size=500):
        overdue = max((today - loan.due_date).days, 0) if loan.due_date and loan.status in ('ACTIVE','DEFAULT') else 0
        ws.append([
            loan.loan_id,
            loan.customer.full_name if loan.customer else '',
            loan.customer.uid       if loan.customer else '',
            loan.product.loan_type  if loan.product  else '',
            loan.branch.name        if loan.branch   else '',
            loan.loan_officer.full_name if loan.loan_officer else '',
            float(loan.principal or 0),
            float(loan.interest_rate or 0),
            float(loan.total_amount or 0),
            float(loan.total_paid   or 0),
            float(loan.balance      or 0),
            loan.status,
            str(loan.disbursed_at.date()) if loan.disbursed_at else '',
            str(loan.due_date) if loan.due_date else '',
            overdue,
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'loans_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    resp = _xl_response(filename)
    resp.write(buf.read())
    return resp


def export_collections_report(queryset=None) -> HttpResponse:
    """Export defaulters / overdue loans to Excel."""
    from apps.loans.models import Loan
    from django.utils import timezone as tz

    today = tz.now().date()
    qs = queryset or Loan.objects.filter(
        status__in=['ACTIVE', 'DEFAULT'],
        due_date__lt=today,
    ).select_related('customer', 'branch', 'loan_officer')

    oxl, has_xl = _workbook()
    if not has_xl:
        from django.http import HttpResponse as HR
        import csv
        resp = HR(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="overdue.csv"'
        w = csv.writer(resp)
        w.writerow(['Loan ID','Customer','Phone','Balance','Days Overdue','Status','Officer'])
        for l in qs:
            days = (today - l.due_date).days if l.due_date else 0
            w.writerow([l.loan_id, l.customer.full_name if l.customer else '',
                        l.customer.phone if l.customer else '',
                        float(l.balance or 0), days, l.status,
                        l.loan_officer.full_name if l.loan_officer else ''])
        return resp

    wb = oxl.Workbook()
    ws = wb.active
    ws.title = 'Overdue Loans'

    headers = ['Loan ID','Customer','Phone','Branch','Balance','Penalty',
               'Days Overdue','Due Date','Status','Officer','Officer Phone']
    ws.append(headers)
    _style_header(ws, 1, oxl)

    from openpyxl.styles import PatternFill, Font
    red_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    red_font = Font(color='DC2626', bold=True)

    for loan in qs.iterator(chunk_size=500):
        days = (today - loan.due_date).days if loan.due_date else 0
        row = ws.append([
            loan.loan_id,
            loan.customer.full_name    if loan.customer      else '',
            loan.customer.phone        if loan.customer      else '',
            loan.branch.name           if loan.branch        else '',
            float(loan.balance         or 0),
            float(loan.penalty_amount  or 0),
            days,
            str(loan.due_date) if loan.due_date else '',
            loan.status,
            loan.loan_officer.full_name  if loan.loan_officer else '',
            getattr(loan.loan_officer, 'phone', '') if loan.loan_officer else '',
        ])
        if days > 30:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill
                cell.font = red_font

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = _xl_response(f'overdue_{timezone.now().strftime("%Y%m%d")}.xlsx')
    resp.write(buf.read())
    return resp


def export_customer_registry(queryset=None) -> HttpResponse:
    """Export customer registry to Excel."""
    from apps.customers.models import Customer

    qs = queryset or Customer.objects.select_related(
        'branch', 'loan_officer'
    ).order_by('-created_at')

    oxl, has_xl = _workbook()
    if not has_xl:
        from apps.customers.bulk import export_customers_csv
        return export_customers_csv(qs)

    wb = oxl.Workbook()
    ws = wb.active
    ws.title = 'Customers'

    headers = ['UID','First Name','Last Name','National ID','Phone','Email',
               'Gender','County','Employment','Employer','Gross Income','Net Salary',
               'Loan Limit','Credit Score','Branch','Loan Officer',
               'Status','Registered','KYC Score']
    ws.append(headers)
    _style_header(ws, 1, oxl)

    for c in qs.iterator(chunk_size=500):
        kyc = sum([bool(c.national_id),bool(c.phone),bool(c.dob),bool(c.gender),
                   bool(c.address),bool(c.county),bool(c.employer),bool(c.next_of_kin)]) * 12
        ws.append([
            c.uid, c.first_name, c.last_name, c.national_id, c.phone,
            c.email or '', c.get_gender_display() if hasattr(c,'get_gender_display') else c.gender,
            c.county or '', c.employment_type or '', c.employer or '',
            float(c.monthly_income or 0), float(c.net_salary or 0),
            float(c.loan_limit or 0), c.credit_score or 0,
            c.branch.name if c.branch else '',
            c.loan_officer.full_name if c.loan_officer else '',
            c.status, str(c.created_at.date()), kyc,
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = _xl_response(f'customers_{timezone.now().strftime("%Y%m%d")}.xlsx')
    resp.write(buf.read())
    return resp
